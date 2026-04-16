import os
import smtplib
import re
import tempfile
import unittest
import uuid
from datetime import date
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from app import create_app, format_time_display, get_weekday_name_ar, saudi_today


class AppTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.db_path = str(Path(tempfile.gettempdir()) / f"csco-test-{uuid.uuid4().hex}.db")
        self.app = create_app(
            {
                "TESTING": True,
                "SECRET_KEY": "test-secret",
                "DATABASE": self.db_path,
                "MAIL_SUPPRESS_SEND": True,
                "OUTBOX": [],
            }
        )
        self.client = self.app.test_client()

    def tearDown(self) -> None:
        if os.path.exists(self.db_path):
            os.remove(self.db_path)

    def login(self, email: str, password: str):
        return self.client.post(
            "/login",
            data={"email": email, "password": password},
            follow_redirects=True,
        )

    def extract_invitation_path(self) -> str:
        outbox = self.app.config["OUTBOX"]
        self.assertTrue(outbox)
        match = re.search(r"http://localhost(?P<path>/complete-registration/[^\s]+)", outbox[-1]["body"])
        self.assertIsNotNone(match)
        return match.group("path")

    def extract_reset_path(self) -> str:
        outbox = self.app.config["OUTBOX"]
        self.assertTrue(outbox)
        match = re.search(r"http://localhost(?P<path>/reset-password/[^\s]+)", outbox[-1]["body"])
        self.assertIsNotNone(match)
        return match.group("path")

    def test_login_page_loads(self) -> None:
        response = self.client.get("/login")
        self.assertEqual(response.status_code, 200)
        self.assertIn("تسجيل الدخول".encode("utf-8"), response.data)
        self.assertIn("إرسال رابط إكمال التسجيل".encode("utf-8"), response.data)

    def test_dashboard_shows_today_name_in_attendance_form(self) -> None:
        response = self.login("ahmed@competitive.local", "Employee@123")
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn("اليوم", page)
        self.assertIn(get_weekday_name_ar(saudi_today()), page)

    def test_admin_can_login_and_export_attendance(self) -> None:
        response = self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.assertEqual(response.status_code, 200)
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "دخول", "recorded_at": "2026-04-17T08:00"},
            follow_redirects=True,
        )
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "خروج", "recorded_at": "2026-04-17T17:00"},
            follow_redirects=True,
        )
        export_response = self.client.get(
            "/attendance/export?employee_id=2&attendance_period=month&attendance_date=2026-04-15"
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            export_response.headers["Content-Type"],
        )
        workbook = load_workbook(BytesIO(export_response.data))
        sheet = workbook.active
        self.assertEqual(sheet.title, "Attendance")
        self.assertEqual(sheet.max_column, 4)
        exported_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertIn(("أحمد علي", "2026-04-17", "08:00 AM", "05:00 PM"), exported_rows)

    def test_dashboard_export_uses_selected_employee(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        response = self.client.get("/dashboard?employee_id=3&attendance_period=day&attendance_date=2026-04-15")
        self.assertEqual(response.status_code, 200)
        self.assertIn('name="employee_id" value="3"'.encode("utf-8"), response.data)
        self.assertIn("/attendance/export".encode("utf-8"), response.data)

    def test_admin_attendance_selection_stays_on_selected_employee(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        response = self.client.get("/dashboard?employee_id=3&attendance_period=day&attendance_date=2026-04-15")
        self.assertEqual(response.status_code, 200)
        self.assertIn('name="user_id"'.encode("utf-8"), response.data)
        self.assertIn('<option value="3" selected>'.encode("utf-8"), response.data)
        self.assertIn("سارة خالد".encode("utf-8"), response.data)

        create_response = self.client.post(
            "/attendance/create",
            data={"user_id": 3, "action": "دخول", "recorded_at": "2026-04-15T08:00"},
            follow_redirects=False,
        )
        self.assertEqual(create_response.status_code, 302)
        self.assertIn("employee_id=3", create_response.headers["Location"])

        export_response = self.client.get(
            "/attendance/export?employee_id=3&attendance_period=month&attendance_date=2026-04-15"
        )
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.data))
        exported_rows = list(workbook.active.iter_rows(min_row=2, values_only=True))
        self.assertIn(("سارة خالد", "2026-04-15", "08:00 AM", "08:00 AM"), exported_rows)
        self.assertNotIn(("أحمد علي", "2026-04-15", "08:00 AM", "08:00 AM"), exported_rows)

    def test_employee_cannot_export_attendance(self) -> None:
        self.login("ahmed@competitive.local", "Employee@123")
        response = self.client.get("/attendance/export", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("متاحة".encode("utf-8"), response.data)
        dashboard_response = self.client.get("/dashboard")
        self.assertNotIn("موظفون مفعلون".encode("utf-8"), dashboard_response.data)

    def test_admin_adds_employee_and_employee_completes_registration(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        response = self.client.post(
            "/employees/create",
            data={"full_name": "موظف دعوة", "email": "invite.employee@competitive.sa"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("بانتظار التفعيل".encode("utf-8"), response.data)
        self.client.post("/logout", follow_redirects=True)

        request_response = self.client.post(
            "/register-request",
            data={"email": "invite.employee@competitive.sa"},
            follow_redirects=True,
        )
        self.assertEqual(request_response.status_code, 200)
        self.assertIn("تم إرسال رابط إكمال التسجيل".encode("utf-8"), request_response.data)
        self.assertEqual(len(self.app.config["OUTBOX"]), 1)

        complete_path = self.extract_invitation_path()
        form_response = self.client.get(complete_path)
        self.assertEqual(form_response.status_code, 200)
        self.assertIn("إنشاء كلمة المرور".encode("utf-8"), form_response.data)

        activate_response = self.client.post(
            complete_path,
            data={"password": "Employee@456", "confirm_password": "Employee@456"},
            follow_redirects=True,
        )
        self.assertEqual(activate_response.status_code, 200)
        self.assertIn("تم تفعيل حسابك بنجاح".encode("utf-8"), activate_response.data)

        login_response = self.login("invite.employee@competitive.sa", "Employee@456")
        self.assertEqual(login_response.status_code, 200)
        self.assertIn("لوحة تشغيل الموارد البشرية".encode("utf-8"), login_response.data)

    def test_register_request_handles_mail_delivery_failure(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/employees/create",
            data={"full_name": "موظف بريد", "email": "mail.fail@competitive.sa"},
            follow_redirects=True,
        )
        self.client.post("/logout", follow_redirects=True)
        with mock.patch("app.send_account_email", side_effect=smtplib.SMTPException("boom")):
            response = self.client.post(
                "/register-request",
                data={"email": "mail.fail@competitive.sa"},
                follow_redirects=True,
            )
        self.assertEqual(response.status_code, 200)
        self.assertIn("تعذر إرسال رابط التفعيل حاليًا".encode("utf-8"), response.data)

    def test_register_request_uses_resend_api_when_configured(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/employees/create",
            data={"full_name": "Ù…ÙˆØ¸Ù Resend", "email": "resend.employee@competitive.sa"},
            follow_redirects=True,
        )
        self.client.post("/logout", follow_redirects=True)
        self.app.config.update(
            MAIL_SUPPRESS_SEND=False,
            MAIL_PROVIDER="resend_api",
            MAIL_FROM="no-reply@mail.competitive.sa",
            RESEND_API_KEY="re_test_key",
            RESEND_API_URL="https://api.resend.com/emails",
        )

        response_mock = mock.MagicMock()
        response_mock.__enter__.return_value = response_mock
        response_mock.__exit__.return_value = None
        response_mock.status = 200
        response_mock.read.return_value = b'{"id":"email_123"}'

        with mock.patch("app.urllib_request.urlopen", return_value=response_mock) as urlopen_mock, mock.patch(
            "app.smtplib.SMTP"
        ) as smtp_class_mock:
            response = self.client.post(
                "/register-request",
                data={"email": "resend.employee@competitive.sa"},
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        urlopen_mock.assert_called_once()
        smtp_class_mock.assert_not_called()

    def test_register_request_rejects_invalid_resend_api_key(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/employees/create",
            data={"full_name": "موظف مفتاح", "email": "bad.key@competitive.sa"},
            follow_redirects=True,
        )
        self.client.post("/logout", follow_redirects=True)
        self.app.config.update(
            MAIL_SUPPRESS_SEND=False,
            MAIL_PROVIDER="resend_api",
            MAIL_FROM="no-reply@mail.competitive.sa",
            RESEND_API_KEY="invalid-token",
        )

        response = self.client.post(
            "/register-request",
            data={"email": "bad.key@competitive.sa"},
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("re_".encode("utf-8"), response.data)

    def test_inactive_employee_cannot_login_before_activation(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/employees/create",
            data={"full_name": "موظف غير مفعل", "email": "inactive.employee@competitive.sa"},
            follow_redirects=True,
        )
        self.client.post("/logout", follow_redirects=True)
        response = self.login("inactive.employee@competitive.sa", "Employee@123")
        self.assertIn("لم يكتمل تفعيله".encode("utf-8"), response.data)

    def test_employee_can_request_reset_password_link(self) -> None:
        response = self.client.post(
            "/reset-password-request",
            data={"email": "ahmed@competitive.local"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("تم إرسال رابط إعادة تعيين كلمة المرور".encode("utf-8"), response.data)

        reset_path = self.extract_reset_path()
        form_response = self.client.get(reset_path)
        self.assertEqual(form_response.status_code, 200)
        self.assertIn("إنشاء كلمة مرور جديدة".encode("utf-8"), form_response.data)

        update_response = self.client.post(
            reset_path,
            data={"password": "Employee@789", "confirm_password": "Employee@789"},
            follow_redirects=True,
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertIn("تم تحديث كلمة المرور بنجاح".encode("utf-8"), update_response.data)

        login_response = self.login("ahmed@competitive.local", "Employee@789")
        self.assertEqual(login_response.status_code, 200)
        self.assertIn("لوحة تشغيل الموارد البشرية".encode("utf-8"), login_response.data)

    def test_payroll_download_is_html(self) -> None:
        self.login("ahmed@competitive.local", "Employee@123")
        response = self.client.get("/payroll/1/download")
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/html", response.headers["Content-Type"])
        self.assertIn("مسير راتب الموظف".encode("utf-8"), response.data)

    def test_admin_can_delete_employee_account_and_related_data(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/employees/create",
            data={"full_name": "موظف للحذف", "email": "delete.me@example.com"},
            follow_redirects=True,
        )
        dashboard_response = self.client.get("/dashboard")
        self.assertIn("delete.me@example.com".encode("utf-8"), dashboard_response.data)

        response = self.client.post("/employees/5/delete", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("تم حذف حساب الموظف".encode("utf-8"), response.data)

    def test_admin_can_delete_attendance_range_for_day(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "دخول", "recorded_at": "2026-04-13T08:00"},
            follow_redirects=True,
        )
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "خروج", "recorded_at": "2026-04-13T17:00"},
            follow_redirects=True,
        )

        response = self.client.post(
            "/attendance/delete-range",
            data={"attendance_period": "day", "attendance_date": "2026-04-13"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("تم حذف".encode("utf-8"), response.data)
        self.assertIn("لا توجد سجلات للفترة المحددة".encode("utf-8"), response.data)

    def test_dashboard_groups_attendance_by_day(self) -> None:
        self.login("aljawhara.ali@competitive.sa", "Admin@123")
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "دخول", "recorded_at": "2026-04-20T08:00"},
            follow_redirects=True,
        )
        self.client.post(
            "/attendance/create",
            data={"user_id": 2, "action": "خروج", "recorded_at": "2026-04-20T17:00"},
            follow_redirects=True,
        )

        response = self.client.get("/dashboard?attendance_period=day&attendance_date=2026-04-20")
        self.assertEqual(response.status_code, 200)
        self.assertIn("سجل الدخول".encode("utf-8"), response.data)
        self.assertIn("سجل الخروج".encode("utf-8"), response.data)
        self.assertIn("08:00 AM".encode("utf-8"), response.data)
        self.assertIn("05:00 PM".encode("utf-8"), response.data)

    def test_attendance_create_overwrites_same_day_same_action(self) -> None:
        self.login("ahmed@competitive.local", "Employee@123")
        self.client.post(
            "/attendance/create",
            data={"action": "دخول", "recorded_at": "2026-04-21T08:00"},
            follow_redirects=True,
        )
        response = self.client.post(
            "/attendance/create",
            data={"action": "دخول", "recorded_at": "2026-04-21T09:30"},
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("تم تحديث آخر حركة لهذا اليوم".encode("utf-8"), response.data)
        dashboard_response = self.client.get("/dashboard?attendance_period=day&attendance_date=2026-04-21")
        self.assertIn("09:30 AM".encode("utf-8"), dashboard_response.data)
        self.assertNotIn("08:00 AM".encode("utf-8"), dashboard_response.data)

    def test_employee_cannot_record_checkout_for_next_day(self) -> None:
        self.login("ahmed@competitive.local", "Employee@123")
        with mock.patch("app.saudi_today", return_value=date(2026, 4, 21)):
            response = self.client.post(
                "/attendance/create",
                data={"action": "\u062e\u0631\u0648\u062c", "recorded_at": "2026-04-22T08:00"},
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("\u064a\u0645\u0643\u0646\u0643 \u062a\u0633\u062c\u064a\u0644 \u0627\u0644\u062e\u0631\u0648\u062c \u0641\u0642\u0637 \u0641\u064a \u0646\u0641\u0633 \u0627\u0644\u064a\u0648\u0645.".encode("utf-8"), response.data)
        dashboard_response = self.client.get("/dashboard?attendance_period=day&attendance_date=2026-04-22")
        self.assertNotIn("08:00 AM".encode("utf-8"), dashboard_response.data)

    def test_format_time_display_returns_time_only(self) -> None:
        self.assertEqual(format_time_display("2026-04-15T20:52"), "08:52 PM")
        self.assertEqual(format_time_display("2026-04-15 20:52:00"), "08:52 PM")
        self.assertEqual(format_time_display("2026-04-15T17:52:00+00:00"), "08:52 PM")
        self.assertEqual(format_time_display(None), "")


if __name__ == "__main__":
    unittest.main()
